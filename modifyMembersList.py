from openpyxl import Workbook
from openpyxl import load_workbook

class Member:
	def __init__(self, name, email, schoolClass):
		self.name = name
		self.email = email
		self.schoolClass = schoolClass

wb1 = load_workbook(filename = 'membersList.xlsx')
wb2 = Workbook()
dest_filename = 'modifiedMemberList.xlsx'

ws1 = wb1.get_sheet_by_name('List')
ws2 = wb2.active
ws2.title = 'modList'

row_count = ws1.max_row

members = []

for row in range(2,row_count):
	col = 2
	firstName = ws1.cell(column = col, row = row).value
	lastName = ws1.cell(column = col+1, row = row).value
	fullName = str(firstName) + ' ' + str(lastName)

	col = 1
	email = ws1.cell(column = col, row = row).value

	col = 4
	schoolClass = ws1.cell(column = col, row = row).value

	newMember = Member(fullName, email, schoolClass)
	members.append(newMember)

row = 5
schoolNames = ['FCRH', 'GSB', 'FCLC']
for member in members:
	if member.name and member.email and member.schoolClass:
		if type(member.schoolClass) == unicode:
			modSchoolClass = str(member.schoolClass.encode('ascii', 'ignore')).upper()
		else:
			modSchoolClass = str(member.schoolClass).upper()

		if '19' in modSchoolClass and '2019' not in modSchoolClass:
			modSchoolClass = modSchoolClass.replace('19', '2019')
		if '18' in modSchoolClass and '2018' not in modSchoolClass:
			modSchoolClass = modSchoolClass.replace('18', '2018')
		if '17' in modSchoolClass and '2017' not in modSchoolClass:
			modSchoolClass = modSchoolClass.replace('17', '2017')
		if '16' in modSchoolClass and '2016' not in modSchoolClass:
			modSchoolClass = modSchoolClass.replace('16', '2016')

		if (modSchoolClass != 'UNKNOWN') and ('2015' not in modSchoolClass) and (modSchoolClass[:4] in schoolNames or modSchoolClass[:3] in schoolNames):
			row = row+1
			_ = ws2.cell(column = 1, row = row, value = member.name)
			_ = ws2.cell(column = 2, row = row, value = member.email)
			_ = ws2.cell(column = 3, row = row, value = modSchoolClass)

wb2.save(filename = dest_filename)

